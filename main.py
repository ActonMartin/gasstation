import os
import json
import time
import requests
import json
import pandas as pd
# 在文件开头添加以下代码
import sys
import io

# 修改标准输出的编码为UTF-8
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# 从配置文件导入API密钥
from config import KEY

class RateLimiter:
    """请求频率限制器

    用于控制API请求频率，确保不超过指定的每秒请求次数
    """
    def __init__(self, max_requests_per_second):
        self.max_requests = max_requests_per_second
        self.requests = []

    def wait_if_needed(self):
        """检查并等待，确保不超过请求频率限制"""
        now = time.time()
        # 清理超过1秒的请求记录
        self.requests = [req_time for req_time in self.requests if now - req_time <= 1]

        if len(self.requests) >= self.max_requests:
            # 计算需要等待的时间
            wait_time = 1 - (now - self.requests[0])
            if wait_time > 0:
                time.sleep(wait_time)
            # 清理旧的请求记录
            self.requests = self.requests[1:]

        self.requests.append(time.time())

# 创建频率限制器实例
rate_limiter = RateLimiter(5)  # 限制每秒5次请求

def get_gas_stations(city, page_index=1, page_size=20):
    """获取指定省份城市的加油站信息

    Args:
        province (str): 省份名称，如'广东'
        city (str): 城市名称，如'深圳'
        page_index (int): 页码，默认1
        page_size (int): 每页返回的数量，默认20

    Returns:
        dict: 加油站信息列表
    """
    # 等待以确保不超过请求频率限制
    rate_limiter.wait_if_needed()

    # 腾讯位置服务API地址
    url = 'https://apis.map.qq.com/ws/place/v1/search'
    
    # 请求参数
    params = {
        'keyword': '加油站',
        'boundary': f'region({city},1)',  # 城市区域搜索，city为城市名，1为层级参数
        'page_size': page_size,
        'page_index': page_index,
        'key': KEY,
        'orderby': '_distance'  # 按距离排序
    }
    
    try:
        response = requests.get(url, params=params)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f'请求失败: {e}', flush=True)
        return None

def save_to_excel(data, filename, pages_data=None):
    """将数据保存到Excel文件

    Args:
        data (dict): API返回的数据
        filename (str): 文件名
        pages_data (list): 每一页的数据列表
    """
    if not data or 'data' not in data:
        return
    
    # 提取所有数据
    all_rows = []
    all_idx = 1
    for item in data['data']:
        station_id = item.get('id', '')
        station_type = item.get('category', '')
        if station_type == '汽车:加油站:中石化':
            station_type_each = '中石化'
        elif station_type == '汽车:加油站:中石油':
            station_type_each = '中石油'
        else:
            station_type_each = '其他'
        row = {
            '序号': str(all_idx),
            'id_tx': station_id,
            '加油站名称': item.get('title', ''),
            '加油站类型': station_type_each,
            '优惠信息表头': '暂无优惠，可上报数据',
            '优惠信息详细': '暂无优惠，可上报数据',
            '加油站地址': item.get('address', ''),
            '加油站电话': item.get('tel', ''),
            '加油站坐标': json.dumps(item.get('location', {}), ensure_ascii=False),
        }
        all_rows.append(row)
        all_idx += 1
    
    # 提取去重后的数据
    unique_rows = []
    unique_idx = 1
    seen_ids = set()  # 记录已出现的加油站ID
    for item in data['data']:
        station_id = item.get('id', '')
        if station_id in seen_ids:
            continue  # 跳过重复ID
        seen_ids.add(station_id)
        station_type = item.get('category', '')
        if station_type == '汽车:加油站:中石化':
            station_type_each = '中石化'
        elif station_type == '汽车:加油站:中石油':
            station_type_each = '中石油'
        else:
            station_type_each = '其他'
        row = {
            '序号': str(unique_idx),
            'id_tx': station_id,
            '加油站名称': item.get('title', ''),
            '加油站类型': station_type_each,
            '优惠信息表头': '暂无优惠，可上报数据',
            '优惠信息详细': '暂无优惠，可上报数据',
            '加油站地址': item.get('address', ''),
            '加油站电话': item.get('tel', ''),
            '加油站坐标': json.dumps(item.get('location', {}), ensure_ascii=False),
        }
        unique_rows.append(row)
        unique_idx += 1
    
    # 创建DataFrame
    df_all = pd.DataFrame(all_rows)
    df_unique = pd.DataFrame(unique_rows)
    
    # 使用ExcelWriter保存到多个工作表
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df_all.to_excel(writer, sheet_name='所有数据', index=False)
        df_unique.to_excel(writer, sheet_name='去重后数据', index=False)
        
        # 保存每一页的数据
        if pages_data:
            for i, page_data in enumerate(pages_data):
                page_idx = i + 1
                page_rows = []
                page_row_idx = 1
                for item in page_data:
                    station_id = item.get('id', '')
                    station_type = item.get('category', '')
                    if station_type == '汽车:加油站:中石化':
                        station_type_each = '中石化'
                    elif station_type == '汽车:加油站:中石油':
                        station_type_each = '中石油'
                    else:
                        station_type_each = '其他'
                    row = {
                        '序号': str(page_row_idx),
                        'id_tx': station_id,
                        '加油站名称': item.get('title', ''),
                        '加油站类型': station_type_each,
                        '优惠信息表头': '暂无优惠，可上报数据',
                        '优惠信息详细': '暂无优惠，可上报数据',
                        '加油站地址': item.get('address', ''),
                        '加油站电话': item.get('tel', ''),
                        '加油站坐标': json.dumps(item.get('location', {}), ensure_ascii=False),
                    }
                    page_rows.append(row)
                    page_row_idx += 1
                
                if page_rows:
                    df_page = pd.DataFrame(page_rows)
                    df_page.to_excel(writer, sheet_name=f'page{page_idx}', index=False)
    
    # 调整列宽和对齐方式
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment
    
    wb = load_workbook(filename)
    
    # 调整所有工作表的列宽和对齐方式
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for column_cells in ws.columns:
            max_length = 0
            column = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
                # 设置单元格水平和垂直居中对齐
                cell.alignment = Alignment(horizontal='center', vertical='center')
            # 调整宽度（中文字符约占2个宽度单位，这里乘以1.2做适当扩展）
            adjusted_width = max_length * 1.2
            # 设置最小宽度为10，避免过窄
            ws.column_dimensions[column].width = adjusted_width if adjusted_width > 10 else 10
    
    wb.save(filename)

def main():
    # 示例省份城市列表
    locations = [
        {'province': '甘肃省', 'city': '兰州市'}
    ]
    
    for location in locations:
        province = location['province']
        city = location['city']
        
        # 创建以省份命名的文件夹
        province_folder = os.path.join(os.getcwd(), province)
        os.makedirs(province_folder, exist_ok=True)
        
        print(f'正在获取{province}{city}的加油站信息...', flush=True)
        # 先获取第一页数据以获取总数量
        page_size = 20
        first_page = get_gas_stations(city, page_index=1, page_size=page_size)
        if not first_page or first_page.get('status') != 0:
            print(f'{city}加油站信息获取失败\n', flush=True)
            continue
        
        count = first_page.get('count', 0)
        print(f'{province}{city}共找到{count}个加油站\n', flush=True)
        total_pages = (count + page_size - 1) // page_size  # 计算总页数
        
        # 获取第一页数据
        first_page_data = first_page.get('data', [])
        
        # 初始化数据列表，创建新的列表避免引用问题
        all_data = first_page_data[:]  # 创建一个新列表
        
        # 保存每一页的数据，同样创建新列表避免引用问题
        pages_data = []
        pages_data.append(first_page_data[:])  # 创建一个新列表
        
        # 循环获取剩余页面
        for page in range(2, total_pages + 1):
            print(f'正在获取第{page}/{total_pages}页数据...', flush=True)
            result = get_gas_stations(city, page_index=page, page_size=page_size)
            if result and result.get('status') == 0:
                current_data = result.get('data', [])
                all_data.extend(current_data)
                pages_data.append(current_data)
            else:
                print(f'第{page}页获取失败，跳过...')
                pages_data.append([])  # 添加空列表表示该页获取失败
        
        if all_data:
            # 保存合并后的数据和每页数据到Excel文件
            filename = f'{province}_{city}_gas_stations.xlsx'
            file_path = os.path.join(province_folder, filename)
            save_to_excel({'data': all_data}, file_path, pages_data)
            print(f'{province}{city}的加油站信息已保存到{file_path}', flush=True)
            
            # 打印统计信息
            print(f'{province}{city}共找到{count}个加油站\n', flush=True)
        else:
            print(f'{city}加油站信息获取失败\n', flush=True)

if __name__ == '__main__':
    main()